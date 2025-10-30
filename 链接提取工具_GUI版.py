"""
链接标题和作者提取工具 - GUI版本
支持文件上传、实时日志显示、结果导出
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import os
import sys
from datetime import datetime
import pandas as pd
import time

# 导入核心提取模块
from extract_links_v4_final import (
    read_excel_with_links,
    get_website_name,
    extract_title_and_author,
    extract_with_playwright_browser,
    extract_platform_info,  # 导入平台特定提取函数
    is_baidu_or_douyin,
    PLAYWRIGHT_AVAILABLE,
    set_gui_log_function  # 导入GUI日志函数设置器
)
import openpyxl
from openpyxl.styles import PatternFill


class LinkExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("链接标题和作者提取工具 v4.6")
        self.root.geometry("1000x750")
        
        # 设置窗口样式和主题
        self.root.resizable(True, True)
        
        # 设置现代化配色
        self.colors = {
            'bg': '#F5F7FA',           # 背景色 - 浅灰蓝
            'primary': '#4A90E2',      # 主色 - 蓝色
            'success': '#5CB85C',      # 成功 - 绿色
            'warning': '#F0AD4E',      # 警告 - 橙色
            'danger': '#D9534F',       # 危险 - 红色
            'info': '#5BC0DE',         # 信息 - 青色
            'dark': '#333333',         # 深色文字
            'light': '#FFFFFF',        # 白色
            'border': '#E0E6ED',       # 边框色
        }
        
        # 设置窗口背景色
        self.root.configure(bg=self.colors['bg'])
        
        # 配置ttk样式
        self.setup_styles()
        
        # 数据存储
        self.input_file = None
        self.output_file = None
        self.is_processing = False
        self.log_queue = queue.Queue()
        
        # 创建界面
        self.create_widgets()
        
        # 启动日志更新线程
        self.update_log_display()
    
    def setup_styles(self):
        """配置Windows优化的现代化样式"""
        style = ttk.Style()
        style.theme_use('clam')  # Windows上效果最好
        
        # 配置LabelFrame - 卡片效果
        style.configure('Card.TLabelframe',
                       background='white',
                       borderwidth=1,
                       relief='solid',
                       bordercolor=self.colors['border'])
        style.configure('Card.TLabelframe.Label',
                       background='white',
                       foreground=self.colors['primary'],
                       font=('Microsoft YaHei UI', 12, 'bold'),
                       padding=(5, 5))
        
        # 配置进度条
        style.configure('Modern.Horizontal.TProgressbar',
                       troughcolor='#E8E8E8',
                       background=self.colors['primary'],
                       borderwidth=0,
                       thickness=22)
        
        # 配置Entry
        style.configure('Modern.TEntry',
                       fieldbackground='white',
                       borderwidth=1,
                       relief='solid')
    
    def create_widgets(self):
        """创建GUI组件"""
        
        # 主容器 - 添加内边距
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=15, pady=15)
        
        # 标题栏 - 美化版
        title_frame = tk.Frame(main_container, bg=self.colors['light'], relief='flat')
        title_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        title_frame.configure(highlightbackground=self.colors['border'], highlightthickness=1)
        
        # 添加内边距
        title_inner = tk.Frame(title_frame, bg=self.colors['light'])
        title_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        title_label = ttk.Label(
            title_inner,
            text="🔗 链接标题和作者提取工具",
            style='Title.TLabel'
        )
        title_label.pack(anchor=tk.W)
        
        version_label = ttk.Label(
            title_inner,
            text="v4.6 - 智能两阶段处理 | 支持今日头条、百度、抖音等主流平台",
            style='Subtitle.TLabel'
        )
        version_label.pack(anchor=tk.W, pady=(5, 0))
        
        # 文件选择区域 - 卡片式设计
        file_frame = ttk.LabelFrame(main_container, text="📁 文件选择", style='Card.TLabelframe', padding="20")
        file_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # 输入文件
        input_label = ttk.Label(file_frame, text="输入文件：", font=('Microsoft YaHei UI', 10, 'bold'))
        input_label.grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        
        input_entry_frame = tk.Frame(file_frame, bg=self.colors['light'])
        input_entry_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        self.input_file_var = tk.StringVar(value="请选择Excel文件...")
        input_entry = ttk.Entry(input_entry_frame, textvariable=self.input_file_var, 
                               width=70, state="readonly", font=('Microsoft YaHei UI', 9))
        input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        input_btn = tk.Button(input_entry_frame, text="📂 浏览文件", command=self.select_input_file,
                             font=('Microsoft YaHei UI', 10), bg=self.colors['primary'], fg='white',
                             activebackground='#3A7BC8', relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        input_btn.pack(side=tk.RIGHT)
        
        # 输出文件
        output_label = ttk.Label(file_frame, text="输出文件：", font=('Microsoft YaHei UI', 10, 'bold'))
        output_label.grid(row=2, column=0, sticky=tk.W, pady=(0, 8))
        
        output_entry_frame = tk.Frame(file_frame, bg=self.colors['light'])
        output_entry_frame.grid(row=3, column=0, sticky=(tk.W, tk.E))
        
        self.output_file_var = tk.StringVar(value="自动生成...")
        output_entry = ttk.Entry(output_entry_frame, textvariable=self.output_file_var,
                                width=70, state="readonly", font=('Microsoft YaHei UI', 9))
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        output_btn = tk.Button(output_entry_frame, text="📂 选择位置", command=self.select_output_file,
                              font=('Microsoft YaHei UI', 10), bg=self.colors['primary'], fg='white',
                              activebackground='#3A7BC8', relief='flat', bd=0, padx=20, pady=8, cursor='hand2')
        output_btn.pack(side=tk.RIGHT)
        
        file_frame.columnconfigure(0, weight=1)
        
        # 控制按钮区域 - Windows原生大按钮
        control_frame = tk.Frame(main_container, bg=self.colors['bg'])
        control_frame.grid(row=2, column=0, pady=(0, 15))
        
        self.start_button = tk.Button(
            control_frame,
            text="▶  开始处理",
            command=self.start_processing,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['primary'],
            fg='white',
            activebackground='#3A7BC8',
            activeforeground='white',
            relief='flat',
            bd=0,
            padx=35,
            pady=14,
            cursor='hand2'
        )
        self.start_button.pack(side=tk.LEFT, padx=6)
        
        self.stop_button = tk.Button(
            control_frame,
            text="⏹  停止处理",
            command=self.stop_processing,
            state=tk.DISABLED,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#95A5A6',
            fg='white',
            disabledforeground='#BDC3C7',
            activebackground='#7F8C8D',
            activeforeground='white',
            relief='flat',
            bd=0,
            padx=35,
            pady=14,
            cursor='hand2'
        )
        self.stop_button.pack(side=tk.LEFT, padx=6)
        
        self.export_button = tk.Button(
            control_frame,
            text="💾  导出结果",
            command=self.export_results,
            state=tk.DISABLED,
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg=self.colors['success'],
            fg='white',
            disabledforeground='#BDC3C7',
            activebackground='#4CAF50',
            activeforeground='white',
            relief='flat',
            bd=0,
            padx=35,
            pady=14,
            cursor='hand2'
        )
        self.export_button.pack(side=tk.LEFT, padx=6)
        
        # 进度条区域 - 美化卡片
        progress_frame = ttk.LabelFrame(main_container, text="📊 处理进度", style='Card.TLabelframe', padding="20")
        progress_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            length=950,
            mode='determinate',
            style='Modern.Horizontal.TProgressbar'
        )
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 12))
        
        self.progress_label = ttk.Label(
            progress_frame, 
            text="等待开始...",
            font=('Microsoft YaHei UI', 10)
        )
        self.progress_label.grid(row=1, column=0, sticky=tk.W, pady=(0, 10))
        
        # 统计信息 - 美化版
        stats_card = tk.Frame(progress_frame, bg=self.colors['light'], relief='flat')
        stats_card.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        stats_card.configure(highlightbackground=self.colors['border'], highlightthickness=1)
        
        stats_inner = tk.Frame(stats_card, bg=self.colors['light'])
        stats_inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=12)
        
        self.stats_label = ttk.Label(
            stats_inner,
            text="总数: 0  |  成功: 0  |  部分: 0  |  失败: 0  |  延迟: 0",
            style='Stats.TLabel'
        )
        self.stats_label.pack()
        
        progress_frame.columnconfigure(0, weight=1)
        
        # 日志显示区域 - 美化卡片
        log_frame = ttk.LabelFrame(main_container, text="📋 处理日志", style='Card.TLabelframe', padding="20")
        log_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 创建滚动文本框 - 更现代的样式
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            width=100,
            height=18,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg='#FAFBFC',
            fg=self.colors['dark'],
            relief='flat',
            borderwidth=1,
            highlightthickness=1,
            highlightbackground=self.colors['border']
        )
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 配置标签颜色 - 更鲜明的配色
        self.log_text.tag_config("success", foreground=self.colors['success'], font=("Consolas", 9, 'bold'))
        self.log_text.tag_config("error", foreground=self.colors['danger'], font=("Consolas", 9, 'bold'))
        self.log_text.tag_config("warning", foreground=self.colors['warning'], font=("Consolas", 9, 'bold'))
        self.log_text.tag_config("info", foreground=self.colors['info'])
        
        # 清空日志按钮 - Windows原生风格
        clear_btn = tk.Button(log_frame, text="🗑️ 清空日志", command=self.clear_log,
                             font=('Microsoft YaHei UI', 9), bg='#F5F5F5', fg=self.colors['dark'],
                             activebackground='#E0E0E0', relief='flat', bd=0, padx=15, pady=6, cursor='hand2')
        clear_btn.grid(row=1, column=0, sticky=tk.E)
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_container.columnconfigure(0, weight=1)
        main_container.rowconfigure(4, weight=1)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # 初始化日志
        self.log("程序已启动", "info")
        if PLAYWRIGHT_AVAILABLE:
            self.log("✅ Playwright已安装，支持百度/抖音/今日头条", "success")
            # 检查浏览器路径（exe环境）
            if getattr(sys, 'frozen', False):
                import os
                base_path = sys._MEIPASS
                
                # 尝试多个可能的浏览器路径
                possible_paths = [
                    os.path.join(base_path, 'playwright_browsers'),
                    os.path.join(base_path, '..', 'playwright_browsers'),
                    os.path.join(os.path.dirname(sys.executable), 'playwright_browsers'),
                ]
                
                found_browser = None
                for path in possible_paths:
                    abs_path = os.path.abspath(path)
                    if os.path.exists(abs_path):
                        found_browser = abs_path
                        break
                
                if found_browser:
                    self.log(f"✅ 使用打包的浏览器（无需额外安装）", "success")
                else:
                    local_playwright_path = os.path.join(os.environ.get('LOCALAPPDATA', ''), 'ms-playwright')
                    if os.path.exists(local_playwright_path):
                        self.log(f"⚠️ 使用本地浏览器: {local_playwright_path}", "warning")
                    else:
                        self.log(f"❌ 未找到浏览器！请运行: 一键安装Chromium浏览器.bat", "error")
        else:
            self.log("⚠️ Playwright未安装，部分功能受限", "warning")
    
    def select_input_file(self):
        """选择输入文件"""
        filename = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if filename:
            self.input_file = filename
            self.input_file_var.set(filename)
            
            # 自动设置输出文件名
            if not self.output_file:
                base_name = os.path.splitext(os.path.basename(filename))[0]
                output_name = f"{base_name}_提取结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                self.output_file = os.path.join(os.path.dirname(filename), output_name)
                self.output_file_var.set(self.output_file)
            
            self.log(f"已选择输入文件: {filename}", "info")
    
    def select_output_file(self):
        """选择输出文件"""
        filename = filedialog.asksaveasfilename(
            title="保存结果文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.output_file = filename
            self.output_file_var.set(filename)
            self.log(f"已设置输出文件: {filename}", "info")
    
    def log(self, message, level="info"):
        """添加日志消息到队列"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_queue.put((timestamp, message, level))
    
    def update_log_display(self):
        """更新日志显示"""
        try:
            while True:
                timestamp, message, level = self.log_queue.get_nowait()
                self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", level)
                self.log_text.see(tk.END)
        except queue.Empty:
            pass
        
        # 每100ms检查一次
        self.root.after(100, self.update_log_display)
    
    def clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
        self.log("日志已清空", "info")
    
    def start_processing(self):
        """开始处理"""
        if not self.input_file:
            messagebox.showerror("错误", "请先选择输入文件！")
            return
        
        if not os.path.exists(self.input_file):
            messagebox.showerror("错误", "输入文件不存在！")
            return
        
        # 设置GUI日志函数（让懂车帝调试日志输出到GUI）
        set_gui_log_function(self.log)
        
        # 禁用开始按钮，启用停止按钮
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.export_button.config(state=tk.DISABLED)
        
        self.is_processing = True
        
        # 在新线程中运行处理
        thread = threading.Thread(target=self.process_links, daemon=True)
        thread.start()
    
    def stop_processing(self):
        """停止处理"""
        self.is_processing = False
        self.log("正在停止处理...", "warning")
    
    def process_links(self):
        """处理链接（在后台线程运行）"""
        try:
            self.log("=" * 60, "info")
            self.log("开始处理链接", "info")
            self.log("=" * 60, "info")
            
            # 读取文件
            self.log(f"正在读取文件: {self.input_file}", "info")
            links = read_excel_with_links(self.input_file)
            total_links = len(links)
            
            self.log(f"找到 {total_links} 个链接", "success")
            
            results = []
            delayed_links = []
            success_count = 0
            partial_count = 0
            failed_count = 0
            
            # 阶段1: 处理普通链接
            self.log("\n" + "=" * 60, "info")
            self.log("【阶段1】处理普通链接（跳过百度/抖音）", "info")
            self.log("=" * 60, "info")
            
            for idx, link_info in enumerate(links, 1):
                if not self.is_processing:
                    self.log("处理已停止", "warning")
                    break
                
                website_name = get_website_name(link_info['url'])
                url_short = link_info['url'][:60]
                
                # 更新进度
                progress = (idx / total_links) * 50  # 阶段1占50%
                self.progress_var.set(progress)
                self.progress_label.config(
                    text=f"阶段1: 处理 {idx}/{total_links} - {website_name}"
                )
                
                # 检查是否是百度或抖音
                if is_baidu_or_douyin(link_info['url']):
                    self.log(f"[{idx}/{total_links}] {website_name} | {url_short}... ⏸️ 延迟处理", "warning")
                    delayed_links.append({
                        'idx': idx,
                        'link_info': link_info,
                        'website_name': website_name
                    })
                    results.append({
                        '原链接': link_info['url'],
                        '网站名': website_name,
                        '作者': '待处理',
                        '标题': '待处理',
                        '状态': 'pending'
                    })
                    continue
                
                # 处理普通链接
                self.log(f"[{idx}/{total_links}] {website_name} | {url_short}...", "info")
                
                info = extract_title_and_author(link_info['url'])
                
                if 'success' in info['status']:
                    self.log(f"  ✅ 成功", "success")
                    success_count += 1
                elif 'partial' in info['status']:
                    self.log(f"  ⚠️ 部分成功", "warning")
                    partial_count += 1
                else:
                    self.log(f"  ❌ 失败: {info['status']}", "error")
                    failed_count += 1
                
                results.append({
                    '原链接': link_info['url'],
                    '网站名': website_name,
                    '作者': info['author'],
                    '标题': info['title'],
                    '状态': info['status']
                })
                
                # 更新统计 - 美化格式
                self.stats_label.config(
                    text=f"总数: {total_links}  |  ✅ 成功: {success_count}  |  ⚠️ 部分: {partial_count}  |  ❌ 失败: {failed_count}  |  ⏸️ 延迟: {len(delayed_links)}"
                )
                
                time.sleep(0.5)
            
            # 阶段2: 处理延迟的链接
            if delayed_links and self.is_processing:
                self.log("\n" + "=" * 60, "info")
                self.log(f"【阶段2】使用Playwright处理延迟的{len(delayed_links)}个链接", "info")
                self.log("=" * 60, "info")
                
                for delayed_idx, delayed_item in enumerate(delayed_links, 1):
                    if not self.is_processing:
                        self.log("处理已停止", "warning")
                        break
                    
                    idx = delayed_item['idx']
                    link_info = delayed_item['link_info']
                    website_name = delayed_item['website_name']
                    url_short = link_info['url'][:60]
                    
                    # 更新进度
                    progress = 50 + (delayed_idx / len(delayed_links)) * 50  # 阶段2占50%
                    self.progress_var.set(progress)
                    self.progress_label.config(
                        text=f"阶段2: 处理 {delayed_idx}/{len(delayed_links)} - {website_name}"
                    )
                    
                    self.log(f"[{delayed_idx}/{len(delayed_links)}] {website_name} | {url_short}...", "info")
                    
                    # 使用平台特定函数（会自动路由到懂车帝、抖音等专用函数）
                    info = extract_platform_info(link_info['url'])
                    
                    if 'success' in info['status']:
                        self.log(f"  ✅ 成功 (Playwright)", "success")
                    elif 'partial' in info['status']:
                        self.log(f"  ⚠️ 部分成功 (Playwright)", "warning")
                    else:
                        self.log(f"  ❌ 失败: {info['status']}", "error")
                    
                    # 更新结果
                    results[idx - 1] = {
                        '原链接': link_info['url'],
                        '网站名': website_name,
                        '作者': info['author'],
                        '标题': info['title'],
                        '状态': info['status']
                    }
                    
                    time.sleep(2)
            
            # 保存结果
            if self.is_processing:
                self.save_results(results)
                self.progress_var.set(100)
                self.progress_label.config(text="处理完成！")
                self.export_button.config(state=tk.NORMAL)
                
                # 统计最终结果
                final_success = sum(1 for r in results if 'success' in r['状态'].lower())
                final_partial = sum(1 for r in results if 'partial' in r['状态'].lower())
                final_failed = len(results) - final_success - final_partial
                
                self.log("\n" + "=" * 60, "info")
                self.log("处理完成！", "success")
                self.log(f"总链接数: {len(results)}", "info")
                self.log(f"完全成功: {final_success} ({final_success/len(results)*100:.1f}%)", "success")
                self.log(f"部分成功: {final_partial} ({final_partial/len(results)*100:.1f}%)", "warning")
                self.log(f"失败: {final_failed} ({final_failed/len(results)*100:.1f}%)", "error")
                self.log(f"结果已保存到: {self.output_file}", "success")
                self.log("=" * 60, "info")
                
                messagebox.showinfo("完成", f"处理完成！\n\n成功: {final_success}\n部分: {final_partial}\n失败: {final_failed}")
        
        except Exception as e:
            self.log(f"处理出错: {str(e)}", "error")
            messagebox.showerror("错误", f"处理失败：{str(e)}")
        
        finally:
            # 恢复按钮状态
            self.start_button.config(state=tk.NORMAL)
            self.stop_button.config(state=tk.DISABLED)
            self.is_processing = False
    
    def save_results(self, results):
        """保存结果到Excel"""
        self.log("正在保存结果...", "info")
        
        df = pd.DataFrame(results)
        df.to_excel(self.output_file, index=False, engine='openpyxl')
        
        # 添加颜色标记
        self.log("添加颜色标记...", "info")
        wb = openpyxl.load_workbook(self.output_file)
        ws = wb.active
        
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        for idx, row_data in enumerate(results, start=2):
            status = row_data['状态'].lower()
            title = row_data['标题']
            author = row_data['作者']
            
            if '404' in status or '404' in title:
                for col in range(1, 6):
                    ws.cell(row=idx, column=col).fill = red_fill
            else:
                if '未找到' in title or '提取失败' in title:
                    ws.cell(row=idx, column=4).fill = yellow_fill
                
                if '未找到' in author or '提取失败' in author:
                    ws.cell(row=idx, column=3).fill = yellow_fill
        
        wb.save(self.output_file)
        self.log("结果已保存", "success")
    
    def export_results(self):
        """导出结果到指定位置"""
        if not self.output_file or not os.path.exists(self.output_file):
            messagebox.showerror("错误", "没有可导出的结果文件！")
            return
        
        export_path = filedialog.asksaveasfilename(
            title="导出结果",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=os.path.basename(self.output_file)
        )
        
        if export_path:
            try:
                import shutil
                shutil.copy2(self.output_file, export_path)
                self.log(f"结果已导出到: {export_path}", "success")
                messagebox.showinfo("成功", f"结果已导出到:\n{export_path}")
            except Exception as e:
                self.log(f"导出失败: {str(e)}", "error")
                messagebox.showerror("错误", f"导出失败：{str(e)}")


def main():
    root = tk.Tk()
    app = LinkExtractorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

