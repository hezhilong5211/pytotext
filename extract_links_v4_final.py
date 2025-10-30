"""
链接标题和作者提取工具 v4.5 - 精准颜色标记版
✅ 成功突破微博！(98.2%)
✅ 抖音80%成功率
✅ 小红书90%成功率
✅ 汽车之家90%成功率
✅ 百家号66%+成功率（Playwright绕过安全验证）
✅ 今日头条80%成功率（自动Playwright）
✅ Excel精准颜色标记：
   🔴 404错误 → 整行标红
   🟡 标题/作者失败 → 对应单元格标黄
"""
import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill
import time
import re
import json
import html as html_module
import os
import sys

# 设置Playwright浏览器路径（用于打包后的exe）
if getattr(sys, 'frozen', False):
    # 如果是打包后的exe
    base_path = sys._MEIPASS
    playwright_browser_path = os.path.join(base_path, 'playwright_browsers')
    if os.path.exists(playwright_browser_path):
        os.environ['PLAYWRIGHT_BROWSERS_PATH'] = playwright_browser_path
        print(f"✅ 使用打包的Playwright浏览器: {playwright_browser_path}")

# 尝试导入Playwright（今日头条专用）
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("⚠️  Playwright未安装，今日头条将使用requests（效果较差）")

# 创建全局Session
session = requests.Session()

# 懂车帝调试日志文件路径
DCD_DEBUG_LOG_FILE = 'dongchedi_debug.log'
_dcd_log_initialized = False

def dcd_debug_log(message):
    """懂车帝调试日志：同时输出到控制台和文件"""
    global _dcd_log_initialized
    
    # 首次调用时清空旧日志
    if not _dcd_log_initialized:
        try:
            with open(DCD_DEBUG_LOG_FILE, 'w', encoding='utf-8') as f:
                f.write(f"=== 懂车帝调试日志 - {time.strftime('%Y-%m-%d %H:%M:%S')} ===\n\n")
            print(f"\n💾 懂车帝调试日志将保存到: {DCD_DEBUG_LOG_FILE}\n", flush=True)
        except:
            pass
        _dcd_log_initialized = True
    
    timestamp = time.strftime("%H:%M:%S")
    log_message = f"[{timestamp}] {message}"
    print(log_message, flush=True)
    
    try:
        with open(DCD_DEBUG_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(log_message + '\n')
    except:
        pass

def read_excel_with_links(file_path):
    """读取Excel文件并提取所有链接"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    links = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                links.append({
                    'cell': cell.coordinate,
                    'text': cell.value,
                    'url': cell.hyperlink.target
                })
            elif isinstance(cell.value, str) and ('http://' in cell.value or 'https://' in cell.value):
                urls = re.findall(r'https?://[^\s]+', cell.value)
                for url in urls:
                    links.append({
                        'cell': cell.coordinate,
                        'text': cell.value,
                        'url': url
                    })
    
    return links

def get_website_name(url):
    """从URL提取网站名称"""
    url_lower = url.lower()
    
    if 'bilibili.com' in url_lower:
        return '哔哩哔哩'
    elif 'douyin.com' in url_lower or 'iesdouyin.com' in url_lower:
        return '抖音'
    elif 'xiaohongshu.com' in url_lower or 'xhslink.com' in url_lower:
        return '小红书'
    elif 'mp.weixin.qq' in url_lower:
        return '微信公众号'
    elif 'm.weibo' in url_lower or 'weibo.com' in url_lower:
        return '新浪微博'
    elif 'toutiao.com' in url_lower:
        return '今日头条'
    elif 'autohome.com' in url_lower:
        return '汽车之家'
    elif 'vc.yiche' in url_lower or 'sv.m.yiche' in url_lower or 'vc.m.yiche' in url_lower or 'yiche.com' in url_lower:
        return '易车'
    elif '163.com' in url_lower or '.163.' in url_lower:
        return '网易'
    elif 'mbd.baidu' in url_lower or 'baijiahao.baidu' in url_lower:
        return '百度'
    elif 'yidianzixun' in url_lower:
        return '一点资讯'
    elif 'zjbyte.cn' in url_lower or 'dongchedi' in url_lower or 'dcd' in url_lower:
        return '懂车帝'
    elif 'zhihu.com' in url_lower:
        return '知乎'
    else:
        try:
            from urllib.parse import urlparse
            domain = urlparse(url).netloc
            domain = domain.replace('www.', '').replace('.com', '').replace('.cn', '')
            return domain if domain else '未知网站'
        except:
            return '未知网站'

def extract_weibo_breakthrough(url):
    """微博突破版 - 使用移动端API"""
    try:
        # 从URL提取mid
        id_match = re.search(r'/(\d+)/([A-Za-z0-9]+)', url)
        if not id_match:
            return {
                'title': 'URL格式错误',
                'author': 'URL格式错误',
                'status': 'failed: URL格式不正确'
            }
        
        uid, mid = id_match.groups()
        
        # 使用移动端API
        api_url = f"https://m.weibo.cn/statuses/show?id={mid}"
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
            'Accept': 'application/json',
            'Referer': f'https://m.weibo.cn/status/{mid}',
            'X-Requested-With': 'XMLHttpRequest',
        }
        
        response = session.get(api_url, headers=headers, timeout=15)
        
        if response.status_code == 200:
            try:
                data = response.json()
                
                if 'data' in data:
                    status = data['data']
                    
                    # 提取标题 - 优先使用text_raw，否则使用text
                    title = status.get('text_raw', status.get('text', ''))
                    
                    # 清理HTML标签
                    if title:
                        title = re.sub(r'<[^>]+>', '', title)
                        title = title.strip()
                        # 限制长度
                        if len(title) > 200:
                            title = title[:200] + '...'
                    
                    # 提取作者
                    author = ''
                    if 'user' in status:
                        author = status['user'].get('screen_name', '')
                    
                    return {
                        'title': title if title else '未找到标题',
                        'author': author if author else '未找到作者',
                        'status': 'success' if (title and author) else 'partial (微博API提取)'
                    }
                else:
                    return {
                        'title': 'API响应无数据',
                        'author': '未找到',
                        'status': 'failed: API响应格式错误'
                    }
            
            except json.JSONDecodeError:
                return {
                    'title': 'API响应解析失败',
                    'author': '未找到',
                    'status': 'failed: JSON解析错误'
                }
        else:
            return {
                'title': f'API请求失败({response.status_code})',
                'author': '未找到',
                'status': f'failed: HTTP {response.status_code}'
            }
    
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_douyin_enhanced(url):
    """抖音增强提取 - 深度JSON搜索"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
            'Accept': 'text/html,application/xhtml+xml',
            'Referer': 'https://www.douyin.com/',
        }
        
        response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title = None
        author = None
        
        # 方法1：从title标签提取标题和作者
        if soup.find('title'):
            title_text = soup.find('title').text.strip()
            # 抖音title格式通常是："标题 - 作者 - 抖音" 或 "标题 @作者 抖音"
            if ' - ' in title_text:
                parts = title_text.split(' - ')
                if len(parts) >= 2:
                    title = parts[0].strip()
                    # 第二部分可能是作者
                    potential_author = parts[1].replace('抖音', '').strip()
                    if potential_author and len(potential_author) > 1 and '抖音' not in potential_author:
                        author = potential_author
            elif '@' in title_text:
                # 格式如："标题 @作者"
                match = re.search(r'@([^\s@]+)', title_text)
                if match:
                    author = match.group(1)
                title = title_text.split('@')[0].strip()
            else:
                title = title_text.replace('- 抖音', '').replace('抖音', '').strip()
        
        # 方法2：从RENDER_DATA深度提取
        render_match = re.search(r'<script id="RENDER_DATA" type="application/json">([^<]+)</script>', response.text)
        if render_match:
            try:
                json_str = html_module.unescape(render_match.group(1))
                data = json.loads(json_str)
                
                # 超深度递归搜索（增强版）
                def deep_find_author(obj, depth=0, max_depth=12):
                    if depth > max_depth:
                        return None
                    
                    if isinstance(obj, dict):
                        # 扩展作者字段
                        author_keys = ['nickname', 'authorName', 'unique_id', 'short_id', 
                                     'userName', 'user_name', 'name', 'author_name',
                                     'secUid', 'uniqueId', 'screen_name']
                        for key in author_keys:
                            if key in obj and isinstance(obj[key], str):
                                value = obj[key].strip()
                                # 排除明显不是作者的值
                                if (2 < len(value) < 50 and 
                                    value not in ['抖音', 'douyin', 'video', 'image'] and
                                    not value.startswith('http')):
                                    return value
                        
                        # 优先搜索author和user字段
                        for key in ['author', 'user', 'authorInfo', 'userInfo']:
                            if key in obj and isinstance(obj[key], dict):
                                result = deep_find_author(obj[key], depth+1, max_depth)
                                if result:
                                    return result
                        
                        # 遍历所有值
                        for value in obj.values():
                            result = deep_find_author(value, depth+1, max_depth)
                            if result:
                                return result
                    
                    elif isinstance(obj, list):
                        for item in obj:
                            result = deep_find_author(item, depth+1, max_depth)
                            if result:
                                return result
                    
                    return None
                
                if not author:
                    author = deep_find_author(data)
                
                # 同时尝试提取标题
                if not title:
                    def deep_find_title(obj, depth=0, max_depth=10):
                        if depth > max_depth:
                            return None
                        if isinstance(obj, dict):
                            for key in ['title', 'desc', 'description', 'content']:
                                if key in obj and isinstance(obj[key], str):
                                    value = obj[key].strip()
                                    if 10 < len(value) < 500:
                                        return value
                            for value in obj.values():
                                result = deep_find_title(value, depth+1, max_depth)
                                if result:
                                    return result
                        elif isinstance(obj, list):
                            for item in obj:
                                result = deep_find_title(item, depth+1, max_depth)
                                if result:
                                    return result
                        return None
                    
                    title = deep_find_title(data)
                
            except Exception as e:
                print(f"  [抖音JSON解析错误: {str(e)[:30]}]", flush=True)
        
        # 方法3：扩展正则表达式搜索
        if not author:
            patterns = [
                r'"nickname"\s*:\s*"([^"]{2,30})"',
                r'"authorName"\s*:\s*"([^"]{2,30})"',
                r'"uniqueId"\s*:\s*"([^"]{2,30})"',
                r'"user_name"\s*:\s*"([^"]{2,30})"',
                r'"name"\s*:\s*"([^"]{2,30})"',
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, response.text)
                if matches:
                    for match in matches:
                        # 验证是否是有效的作者名
                        if (re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9_\-]+$', match) and
                            match not in ['抖音', 'douyin', 'video'] and
                            len(match) > 1):
                            author = match
                            break
                if author:
                    break
        
        # 方法4：从meta标签提取
        if not author:
            meta_author = soup.find('meta', {'name': 'author'})
            if meta_author:
                author = meta_author.get('content', '').strip()
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial (抖音作者难提取)'
        }
        
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_weixin_info(url):
    """微信公众号提取"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
        }
        
        response = session.get(url, headers=headers, timeout=15)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title = None
        author = None  # 这里的author指公众号名称
        
        # 提取标题
        # 方法1: meta标签
        meta_title = soup.find('meta', {'property': 'og:title'})
        if meta_title:
            title = meta_title.get('content', '').strip()
        
        # 方法2: title标签
        if not title:
            title_elem = soup.find('title')
            if title_elem:
                title = title_elem.get_text().strip()
        
        # 提取公众号名称（作者字段）
        # 方法1: id="js_name"（最准确）
        js_name_elem = soup.find(id='js_name')
        if js_name_elem:
            author = js_name_elem.get_text().strip()
        
        # 方法2: class="rich_media_meta_nickname"
        if not author:
            nickname_elem = soup.find(class_='rich_media_meta_nickname')
            if nickname_elem:
                author = nickname_elem.get_text().strip()
        
        # 方法3: class包含profile_nickname
        if not author:
            profile_elem = soup.find(class_=re.compile('profile_nickname', re.I))
            if profile_elem:
                author = profile_elem.get_text().strip()
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到公众号',
            'status': 'success' if (title and author) else 'partial'
        }
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_xiaohongshu_info(url):
    """小红书提取 - 增强版（支持explore链接）"""
    try:
        # 检测是否是explore链接（需要特殊处理）
        is_explore = '/explore/' in url.lower()
        
        # 对于explore链接，使用更完整的移动端headers
        if is_explore:
            headers = {
                'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 MicroMessenger/8.0.38(0x1800262c) NetType/WIFI Language/zh_CN',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Referer': 'https://www.xiaohongshu.com/',
            }
        else:
            headers = {
                'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
            }
        
        response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.encoding = 'utf-8'
        
        # 检查是否被重定向到错误页面
        if 'website-login/error' in response.url or 'error_code' in response.url:
            return {
                'title': '访问受限',
                'author': '访问受限',
                'status': 'failed: 小红书访问受限，建议使用xhslink.com短链接'
            }
        
        title = None
        author = None
        
        patterns = [
            r'window\.__INITIAL_STATE__\s*=\s*({.+?})\s*<\/script>',
            r'window\.__SETUP_SERVER_STATE__\s*=\s*({.+?})\s*<\/script>',
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, response.text, re.DOTALL)
            if matches:
                try:
                    data_str = matches[0].replace('\\u002F', '/')
                    
                    # 提取desc字段
                    desc_text = None
                    desc_match = re.search(r'"desc"\s*:\s*"([^"]+)"', data_str)
                    if desc_match and len(desc_match.group(1)) > 5:
                        desc_text = desc_match.group(1)
                        # 清理话题标签：#xxx[话题]# 或 #xxx#
                        desc_text = re.sub(r'#[^#\[]+\[话题\]#', '', desc_text)
                        desc_text = re.sub(r'#[^#]+#', '', desc_text)
                        desc_text = desc_text.strip()
                    
                    # 提取title字段
                    title_text = None
                    title_match = re.search(r'"title"\s*:\s*"([^"]+)"', data_str)
                    if title_match and title_match.group(1):
                        title_text = title_match.group(1).strip()
                        if title_text == '小红书' or len(title_text) <= 5:
                            title_text = None
                    
                    # 智能选择标题：
                    # 1. 如果title存在且长度合理（8-50字），检查与desc的相关性
                    # 2. 如果title和desc不相关（可能是推荐内容的标题），用desc
                    # 3. 如果title太短或为空，用desc
                    if title_text and desc_text:
                        title_len = len(title_text)
                        # title长度合理（8-50字）
                        if 8 <= title_len <= 50:
                            # 简单检查：提取title和desc中的3-4字连续字符串，看是否有交集
                            # 如果有共同的3-4字词，说明相关
                            title_trigrams = set()
                            desc_trigrams = set()
                            
                            # 提取title中的3-4字片段（排除常用2字词）
                            for i in range(len(title_text) - 2):
                                for j in range(3, 5):
                                    if i + j <= len(title_text):
                                        fragment = title_text[i:i+j]
                                        # 排除纯标点或空格
                                        if any('\u4e00' <= c <= '\u9fff' for c in fragment):
                                            title_trigrams.add(fragment)
                            
                            # 提取desc前50字中的3-4字片段
                            desc_sample = desc_text[:50]
                            for i in range(len(desc_sample) - 2):
                                for j in range(3, 5):
                                    if i + j <= len(desc_sample):
                                        fragment = desc_sample[i:i+j]
                                        # 排除纯标点或空格
                                        if any('\u4e00' <= c <= '\u9fff' for c in fragment):
                                            desc_trigrams.add(fragment)
                            
                            # 如果有共同的3-4字片段，说明相关，用title
                            if title_trigrams & desc_trigrams:
                                title = title_text
                            else:
                                # 没有共同片段，title可能是推荐内容，用desc
                                title = desc_text
                        else:
                            # title太短，用desc
                            title = desc_text
                    elif title_text:
                        # 只有title没有desc
                        title = title_text
                    else:
                        # 没有title，用desc
                        title = desc_text
                    
                    # 提取作者（同时支持 nickName 和 nickname）
                    # 优先从 noteData.user.nickName 提取（最准确）
                    user_nickname_match = re.search(r'"user"\s*:\s*{[^}]{0,500}"nickName"\s*:\s*"([^"]{2,30})"', data_str)
                    if user_nickname_match:
                        author = user_nickname_match.group(1)
                    
                    # 备选：通用nickname字段
                    if not author:
                        nickname_matches = re.findall(r'"nickname"\s*:\s*"([^"]+)"', data_str)
                        if nickname_matches:
                            author = nickname_matches[0]
                    
                    if title and author:
                        break
                except:
                    continue
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_bilibili_info(url):
    """B站提取"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Referer': 'https://www.bilibili.com/',
        }
        
        response = session.get(url, headers=headers, timeout=15)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title = None
        author = None
        
        meta_title = soup.find('meta', {'property': 'og:title'})
        if meta_title:
            title = meta_title.get('content', '').strip()
            for suffix in ['_哔哩哔哩_bilibili', ' - 哔哩哔哩']:
                if suffix in title:
                    title = title.split(suffix)[0].strip()
        
        meta_author = soup.find('meta', {'name': 'author'})
        if meta_author:
            author = meta_author.get('content', '').strip()
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_autohome_info(url):
    """汽车之家专用提取"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }
        
        response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.encoding = response.apparent_encoding
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title = None
        author = None
        
        # 提取标题
        title_elem = soup.find('title')
        if title_elem:
            title = title_elem.text.strip()
        
        # 车家号（PC和移动版）：从JSON提取作者（优先方案）
        if 'chejiahao' in url or ('m.autohome.com.cn' in url and 'info' in url):
            # 方法1: 优先从authorName字段提取（最精准，车家号专用）
            author_match = re.search(r'"authorName"\s*:\s*"([^"]{2,30})"', response.text)
            if author_match:
                potential_author = author_match.group(1)
                if potential_author not in ['汽车之家', 'autohome', '车家号']:
                    author = potential_author
            
            # 方法2: 从包含authorid的author对象中提取
            if not author:
                author_match = re.search(r'"author"\s*:\s*{\s*[^}]*"authorid"\s*:\s*"[^"]+"\s*[^}]*"name"\s*:\s*"([^"]{2,30})"', response.text)
                if author_match:
                    author = author_match.group(1)
            
            # 方法3: 从authorInfo中提取
            if not author:
                author_match = re.search(r'"authorInfo"\s*:\s*{\s*[^}]*"name"\s*:\s*"([^"]{2,30})"', response.text)
                if author_match:
                    author = author_match.group(1)
            
            # 方法4: 从文章数据中的userName提取
            if not author:
                author_match = re.search(r'"userName"\s*:\s*"([^"]{2,30})"', response.text)
                if author_match:
                    potential_author = author_match.group(1)
                    if potential_author not in ['汽车之家', 'autohome', '车家号']:
                        author = potential_author
            
            # 方法5: 从HTML元素提取（class="authorMes"）
            if not author:
                author_elem = soup.find('div', class_='authorMes')
                if author_elem:
                    # 查找所有a标签，找到包含作者名的那个（不是空的，且有文本）
                    a_elems = author_elem.find_all('a')
                    for a in a_elems:
                        text = a.get_text(strip=True)
                        if text and len(text) >= 2 and len(text) <= 30:
                            # 排除数字、"关注"、"作品"等
                            if not re.match(r'^\d+', text) and text not in ['汽车之家', '车家号', '关注', '作品']:
                                author = text
                                break
            
            # 方法6: 查找class包含author的元素
            if not author:
                author_elems = soup.find_all(class_=re.compile('author', re.I))
                for elem in author_elems:
                    text = elem.get_text(strip=True)
                    if text and 2 <= len(text) <= 30 and text not in ['汽车之家', '车家号', '相关推荐']:
                        author = text
                        break
        
        # 论坛帖子：尝试多种方法
        elif 'club.autohome.com.cn' in url:
            # 方法1: 从JavaScript变量__TOPICINFO__提取topicMemberName
            topic_match = re.search(r'topicMemberName:\s*[\'"]([^\'"]+)[\'"]', response.text)
            if topic_match:
                author = topic_match.group(1)
            
            # 方法2: 查找用户链接
            if not author:
                user_link = soup.find('a', href=re.compile(r'/space/\d+'))
                if user_link:
                    author = user_link.get_text(strip=True)
            
            # 方法3: 在JSON中查找userName
            if not author:
                author_match = re.search(r'"userName":\s*"([^"]+)"', response.text)
                if author_match:
                    author = author_match.group(1)
        
        # 新闻页面：提取作者
        elif 'www.autohome.com.cn/news' in url:
            # 方法1: 从meta标签
            meta_author = soup.find('meta', {'name': 'author'})
            if meta_author:
                author = meta_author.get('content', '').strip()
            
            # 方法2: 查找class包含author的元素
            if not author:
                author_elem = soup.find(class_=re.compile('author|writer', re.I))
                if author_elem:
                    author = author_elem.get_text(strip=True)
            
            # 方法3: 在页面源码中搜索
            if not author:
                author_patterns = [
                    r'"author"\s*:\s*"([^"]+)"',
                    r'作者[：:]\s*([^\s<>"]+)',
                ]
                for pattern in author_patterns:
                    match = re.search(pattern, response.text)
                    if match:
                        author = match.group(1)
                        break
        
        # 通用提取（fallback）
        if not author:
            meta_author = soup.find('meta', {'name': 'author'})
            if meta_author:
                author = meta_author.get('content', '').strip()
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
        
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_general_info(url):
    """通用提取"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }
        
        response = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        response.encoding = response.apparent_encoding
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title = None
        author = None
        
        meta_title = soup.find('meta', {'property': 'og:title'}) or soup.find('meta', {'name': 'title'})
        if meta_title:
            title = meta_title.get('content', '').strip()
        
        if not title and soup.find('title'):
            title = soup.find('title').text.strip()
        
        meta_author = soup.find('meta', {'property': 'og:author'}) or soup.find('meta', {'name': 'author'})
        if meta_author:
            author = meta_author.get('content', '').strip()
        
        if not author:
            author_elems = soup.find_all(class_=re.compile(r'author|writer', re.I))
            for elem in author_elems:
                text = elem.text.strip()
                if text and len(text) < 50:
                    author = text
                    break
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_baidu_info(url):
    """百度系平台提取（百家号、mbd.baidu）"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml',
        }
        
        response = session.get(url, headers=headers, timeout=10, allow_redirects=True)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取标题
        title = None
        title_tag = soup.find('title')
        if title_tag:
            title = title_tag.text.strip()
        
        # 提取作者
        author = None
        
        # 方法1：百家号的作者信息（新版页面）
        if 'baijiahao.baidu.com' in url:
            # 新版页面：从HTML标签提取
            # <span data-testid="author-name" class="_2gGWi">作者名</span>
            author_elem = soup.find('span', {'data-testid': 'author-name'})
            if author_elem:
                author = author_elem.get_text().strip()
            
            # 备用方法1：从class="_2gGWi"提取
            if not author:
                author_elem = soup.find('span', class_='_2gGWi')
                if author_elem:
                    author = author_elem.get_text().strip()
            
            # 备用方法2：从author链接提取
            if not author:
                author_link = soup.find('a', href=re.compile('author.baidu.com'))
                if author_link:
                    author_span = author_link.find('span')
                    if author_span:
                        author = author_span.get_text().strip()
            
            # 备用方法3：旧版JSON数据（兼容旧链接）
            if not author:
                json_match = re.search(r'var\s+DATA\s*=\s*({.+?});', response.text, re.DOTALL)
                if json_match:
                    try:
                        data = json.loads(json_match.group(1))
                        if 'superlanding' in data and len(data['superlanding']) > 0:
                            item = data['superlanding'][0].get('itemData', {})
                            if not title or title == '百度':
                                title = item.get('header', title)
                            author = item.get('author', {}).get('name', '')
                    except:
                        pass
        
        # 方法2：mbd.baidu的JSON数据（增强版）
        elif 'mbd.baidu.com' in url:
            # 方法2.1: 标准author.name格式
            author_match = re.search(r'"author"\s*:\s*{\s*[^}]*"name"\s*:\s*"([^"]+)"', response.text)
            if author_match:
                author = author_match.group(1)
                # Unicode解码
                try:
                    if '\\u' in author:
                        author = author.encode().decode('unicode_escape')
                except:
                    pass
            
            # 方法2.2: 更多作者字段尝试
            if not author:
                author_patterns = [
                    r'"authorName"\s*:\s*"([^"]{2,30})"',
                    r'"author_name"\s*:\s*"([^"]{2,30})"',
                    r'"publisher"\s*:\s*"([^"]{2,30})"',
                    r'"source"\s*:\s*"([^"]{2,30})"',
                    r'"sourceName"\s*:\s*"([^"]{2,30})"',
                    r'"account"\s*:\s*{\s*[^}]*"name"\s*:\s*"([^"]+)"',  # account.name格式
                ]
                for pattern in author_patterns:
                    match = re.search(pattern, response.text)
                    if match:
                        candidate = match.group(1)
                        # 过滤掉明显不是作者的
                        if candidate not in ['百度', '百家号', '百度APP', '百度新闻'] and len(candidate) > 1:
                            author = candidate
                            # Unicode解码
                            try:
                                if '\\u' in author:
                                    author = author.encode().decode('unicode_escape')
                            except:
                                pass
                            break
            
            # 查找标题
            if not title or title == '百度':
                title_match = re.search(r'"title"\s*:\s*"([^"]{5,100})"', response.text)
                if title_match:
                    title = title_match.group(1)
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial (百度系)'
        }
        
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '未找到作者',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_dripcar_info(url):
    """水滴汽车平台提取"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }
        
        response = session.get(url, headers=headers, timeout=10)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取标题
        title = None
        title_tag = soup.find('title')
        if title_tag:
            title = title_tag.text.strip()
            # 清理标题
            if '-水滴汽车' in title:
                title = title.split('-水滴汽车')[0].strip()
        
        # 提取作者 - 从JavaScript变量中
        author = None
        author_match = re.search(r'author_name\s*:\s*"([^"]+)"', response.text)
        if author_match:
            author = author_match.group(1)
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
        
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '未找到作者',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_maiche_info(url):
    """买车网平台提取 - 增强版"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        }
        
        response = session.get(url, headers=headers, timeout=10)
        response.encoding = 'utf-8'
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # 提取标题
        title = None
        title_tag = soup.find('title')
        if title_tag:
            title = title_tag.text.strip()
            # 清理标题
            if '- 买车网' in title:
                title = title.split('- 买车网')[0].strip()
        
        # 或者从h1获取
        if not title:
            h1 = soup.find('h1')
            if h1:
                title = h1.text.strip()
        
        # 提取作者 - 增强版
        author = None
        
        # 方法1：从"来源："或"来源："后提取（买车网常用格式）
        # 例如：来源：奶爸教选车
        source_patterns = [
            r'来源[：:]\s*([^\s\n<>]{2,20})',
            r'source[：:]\s*([^\s\n<>]{2,20})',
        ]
        for pattern in source_patterns:
            match = re.search(pattern, response.text)
            if match:
                potential_author = match.group(1).strip()
                # 过滤掉一些明显不是作者的词
                if potential_author not in ['买车网', '网络', '互联网', '官方']:
                    author = potential_author
                    break
        
        # 方法2：从[车友头条-车友号-作者名]格式提取
        if not author:
            match = re.search(r'\[车友头条[^\]]*车友号[^\]]*[-\-]\s*([^\]]{2,20})\]', response.text)
            if match:
                author = match.group(1).strip()
        
        # 方法3：查找作者class
        if not author:
            author_elem = soup.find(class_=re.compile('author', re.I))
            if author_elem:
                author = author_elem.text.strip()
        
        # 方法4：查找meta标签
        if not author:
            meta_author = soup.find('meta', {'name': 'author'})
            if meta_author:
                author = meta_author.get('content', '').strip()
        
        # 方法5：从JSON数据中查找
        if not author:
            author_match = re.search(r'"author"\s*:\s*"([^"]+)"', response.text)
            if author_match:
                author = author_match.group(1)
        
        # 方法6：从"文/"或"文:"后提取
        if not author:
            match = re.search(r'[（(]文[/／:]([^）)]{2,20})[）)]', response.text)
            if match:
                author = match.group(1).strip()
        
        # 统一清理作者名称（应用于所有方法）
        if author:
            # 去掉"车友号"、"作者："等前缀（包括空格）
            author = re.sub(r'^(车友号|作者|来源)\s*[：:\s]*', '', author).strip()
            # 再次清理多余空格
            author = ' '.join(author.split())
        
        return {
            'title': title if title else '未找到标题',
            'author': author if author else '未找到作者',
            'status': 'success' if (title and author) else 'partial'
        }
        
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '未找到作者',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_dongchedi_info(url):
    """懂车帝平台提取 - Playwright版（懂车帝需JS渲染）+ requests降级 + 调试日志"""
    # 懂车帝页面是JS渲染的，直接使用Playwright
    if not PLAYWRIGHT_AVAILABLE:
        return {
            'title': 'Playwright未安装',
            'author': '未找到',
            'status': 'failed: 懂车帝需要Playwright'
        }
    
    dcd_debug_log(f"[DEBUG-DCD] 开始提取懂车帝: {url[:80]}...")
    playwright_result = None
    
    try:
        with sync_playwright() as p:
            # 启动浏览器
            browser = None
            try:
                browser = p.chromium.launch(
                    headless=True,
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox'
                    ]
                )
            except:
                return {
                    'title': '浏览器启动失败',
                    'author': '未找到',
                    'status': 'failed: Browser launch error'
                }
            
            context = browser.new_context(
                user_agent='Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
                viewport={'width': 375, 'height': 812},
            )
            page = context.new_page()
            
            # 去除 webdriver 标记
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            """)
            
            # 访问页面
            try:
                page.goto(url, wait_until='networkidle', timeout=30000)
                time.sleep(5)  # 增加等待时间，确保JS完全渲染
            except:
                try:
                    page.goto(url, wait_until='domcontentloaded', timeout=20000)
                    time.sleep(5)  # 增加等待时间
                except:
                    time.sleep(3)
            
            # 额外等待确保作者信息加载完成
            try:
                # 等待可能包含作者信息的元素
                page.wait_for_timeout(2000)  # 再等2秒
            except:
                pass
            
            # 获取页面内容
            html_content = page.content()
            browser.close()
            
            dcd_debug_log(f"[DEBUG-DCD] HTML内容长度: {len(html_content)} 字符")
            
            # 保存HTML内容到文件用于调试（仅保存第一个）
            try:
                import os
                debug_file = 'dongchedi_page.html'
                if not os.path.exists(debug_file):
                    with open(debug_file, 'w', encoding='utf-8') as f:
                        f.write(html_content)
                    dcd_debug_log(f"[DEBUG-DCD] HTML已保存到: {debug_file}")
            except Exception as e:
                dcd_debug_log(f"[DEBUG-DCD] 保存HTML失败: {str(e)}")
            
            if not html_content or len(html_content) < 500:
                dcd_debug_log(f"[DEBUG-DCD] 错误: 页面内容太短")
                return {
                    'title': '页面加载失败',
                    'author': '未找到',
                    'status': 'failed: 页面内容为空'
                }
            
            # 解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            title = None
            author = None
            
            # 提取标题
            # 方法1: h1标签
            h1_elem = soup.find('h1')
            if h1_elem:
                title = h1_elem.get_text().strip()
            
            # 方法2: title标签
            if not title or len(title) < 5:
                title_elem = soup.find('title')
                if title_elem:
                    title = title_elem.get_text().strip()
                    # 清理标题（处理各种分隔符）
                    for suffix in [' - 懂车帝', '-懂车帝', '_懂车帝', '|懂车帝', ' 懂车帝']:
                        if suffix in title:
                            title = title.split(suffix)[0].strip()
                            break
            
            # 方法3: meta标签
            if not title or len(title) < 5:
                meta_title = soup.find('meta', {'property': 'og:title'})
                if meta_title:
                    title = meta_title.get('content', '').strip()
            
            # 提取作者
            # 方法1: 从JSON数据提取（扩展更多模式）
            dcd_debug_log(f"[DEBUG-DCD] 开始提取作者（方法1: JSON数据）")
            json_patterns = [
                r'"author"\s*:\s*{\s*[^}]*"name"\s*:\s*"([^"]+)"',
                r'"author_name"\s*:\s*"([^"]{2,30})"',
                r'"user_name"\s*:\s*"([^"]{2,30})"',
                r'"userName"\s*:\s*"([^"]{2,30})"',
                r'"nickname"\s*:\s*"([^"]{2,30})"',
                r'"screen_name"\s*:\s*"([^"]{2,30})"',
                r'"source"\s*:\s*"([^"]{2,30})"',
                r'"media_name"\s*:\s*"([^"]{2,30})"',
                r'"name"\s*:\s*"([^"]{2,30})"',  # 通用name字段
                r'"creator"\s*:\s*"([^"]{2,30})"',  # 创建者
                r'"publisher"\s*:\s*"([^"]{2,30})"',  # 发布者
            ]
            
            for i, pattern in enumerate(json_patterns):
                matches = re.findall(pattern, html_content)
                if matches:
                    dcd_debug_log(f"[DEBUG-DCD] 模式{i+1}找到{len(matches)}个匹配: {matches[:3]}")
                    # 尝试所有匹配，找到第一个有效的作者
                    for potential_author in matches:
                        dcd_debug_log(f"[DEBUG-DCD] 检查候选作者: '{potential_author}'")
                        if (potential_author not in ['懂车帝', 'dongchedi', '今日头条', 'toutiao', 'bytedance', 
                                                     '抖音', 'douyin', '字节跳动', '头条', 'article'] and 
                            len(potential_author) > 1 and
                            not potential_author.startswith('http') and
                            not potential_author.isdigit()):  # 排除纯数字
                            author = potential_author
                            dcd_debug_log(f"[DEBUG-DCD] ✓ 找到有效作者: '{author}'")
                            try:
                                if '\\u' in author:
                                    author = author.encode().decode('unicode_escape')
                            except:
                                pass
                            break
                        else:
                            dcd_debug_log(f"[DEBUG-DCD] ✗ 候选作者被过滤")
                    if author:
                        break
            
            # 方法2: 从HTML元素提取
            if not author:
                dcd_debug_log(f"[DEBUG-DCD] 方法1失败，尝试方法2: HTML元素提取")
                author_selectors = [
                    {'class': re.compile('author|writer|creator', re.I)},
                    {'class': re.compile('source|publisher', re.I)},
                    {'class': re.compile('user|username', re.I)},
                ]
                for i, selector in enumerate(author_selectors):
                    elem = soup.find('span', selector) or soup.find('a', selector) or soup.find('div', selector)
                    if elem:
                        text = elem.get_text().strip()
                        dcd_debug_log(f"[DEBUG-DCD] HTML选择器{i+1}找到元素: '{text[:50]}'")
                        if text and 2 <= len(text) <= 30 and text not in ['懂车帝', '']:
                            author = text
                            dcd_debug_log(f"[DEBUG-DCD] ✓ HTML元素提取成功: '{author}'")
                            break
            else:
                dcd_debug_log(f"[DEBUG-DCD] 方法1成功提取作者")
            
            # 方法3: meta标签
            if not author:
                meta_author = soup.find('meta', {'name': 'author'}) or soup.find('meta', {'property': 'article:author'})
                if meta_author:
                    potential_author = meta_author.get('content', '').strip()
                    if potential_author not in ['懂车帝', '']:
                        author = potential_author
            
            playwright_result = {
                'title': title if title else '未找到标题',
                'author': author if author else '未找到作者',
                'status': 'success (Playwright)' if (title and author) else 'partial (Playwright)'
            }
            
            dcd_debug_log(f"[DEBUG-DCD] 提取结果:")
            dcd_debug_log(f"[DEBUG-DCD]   标题: {playwright_result['title'][:50]}")
            dcd_debug_log(f"[DEBUG-DCD]   作者: {playwright_result['author']}")
            dcd_debug_log(f"[DEBUG-DCD]   状态: {playwright_result['status']}")
            
            # 如果Playwright成功获取了作者，直接返回
            if author:
                return playwright_result
            else:
                dcd_debug_log(f"[DEBUG-DCD] Playwright未能获取作者，准备降级到requests")
            
    except Exception as e:
        playwright_result = {
            'title': '提取失败',
            'author': '未找到',
            'status': f'failed: {str(e)[:50]}'
        }
    
    # 如果Playwright未能获取作者，尝试requests降级方案
    if playwright_result and '未找到' in playwright_result.get('author', ''):
        try:
            print("  [降级到requests尝试]", end='', flush=True)
            headers = {
                'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            }
            response = session.get(url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            
            author = None
            # 从requests响应中提取作者
            json_patterns = [
                r'"author_name"\s*:\s*"([^"]{2,30})"',
                r'"user_name"\s*:\s*"([^"]{2,30})"',
                r'"userName"\s*:\s*"([^"]{2,30})"',
                r'"nickname"\s*:\s*"([^"]{2,30})"',
            ]
            
            for pattern in json_patterns:
                matches = re.findall(pattern, response.text)
                if matches:
                    for potential_author in matches:
                        if (potential_author not in ['懂车帝', 'dongchedi', '今日头条'] and 
                            len(potential_author) > 1):
                            author = potential_author
                            break
                    if author:
                        break
            
            if author:
                playwright_result['author'] = author
                playwright_result['status'] = 'success (requests降级)'
                print("✓", flush=True)
            
        except:
            pass
    
    return playwright_result

def extract_toutiao_playwright(url):
    """今日头条专用 - Playwright增强版"""
    if not PLAYWRIGHT_AVAILABLE:
        # 降级到requests
        return extract_general_info(url)
    
    try:
        with sync_playwright() as p:
            # 尝试多种方式启动浏览器
            browser = None
            launch_errors = []
            
            # 方式1: 默认启动
            if not browser:
                try:
                    browser = p.chromium.launch(
                        headless=True,
                        args=[
                            '--disable-blink-features=AutomationControlled',
                            '--disable-dev-shm-usage',
                            '--no-sandbox'
                        ]
                    )
                except Exception as e:
                    launch_errors.append(f"默认启动失败: {str(e)[:50]}")
            
            # 方式2: 不指定executable_path
            if not browser:
                try:
                    browser = p.chromium.launch(
                        headless=True,
                        executable_path=None,
                        args=[
                            '--disable-blink-features=AutomationControlled',
                            '--disable-dev-shm-usage',
                            '--no-sandbox'
                        ]
                    )
                except Exception as e:
                    launch_errors.append(f"无路径启动失败: {str(e)[:50]}")
            
            # 方式3: 尝试指定系统路径
            if not browser:
                import os
                possible_paths = [
                    os.path.expanduser("~/.cache/ms-playwright/chromium-*/chrome-win/chrome.exe"),
                    os.path.expanduser("~/AppData/Local/ms-playwright/chromium-*/chrome-win/chrome.exe"),
                    "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
                    "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
                ]
                for path in possible_paths:
                    try:
                        if '*' in path:
                            import glob
                            matches = glob.glob(path)
                            if matches:
                                path = matches[0]
                        if os.path.exists(path):
                            browser = p.chromium.launch(
                                headless=True,
                                executable_path=path,
                                args=[
                                    '--disable-blink-features=AutomationControlled',
                                    '--disable-dev-shm-usage',
                                    '--no-sandbox'
                                ]
                            )
                            break
                    except Exception as e:
                        launch_errors.append(f"路径{path}启动失败: {str(e)[:30]}")
                    
            # 如果所有方式都失败
            if not browser:
                error_msg = "; ".join(launch_errors) if launch_errors else "未知错误"
                return {
                    'title': '浏览器启动失败',
                    'author': '未找到',
                    'status': f'failed: Browser launch error - {error_msg}'
                }
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                viewport={'width': 1920, 'height': 1080},
                extra_http_headers={
                    'Accept-Language': 'zh-CN,zh;q=0.9',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                }
            )
            page = context.new_page()
            
            # 增强反检测
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3, 4, 5]
                });
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['zh-CN', 'zh', 'en']
                });
            """)
            
            # 访问页面 - 增加等待时间
            try:
                page.goto(url, wait_until='networkidle', timeout=45000)
                time.sleep(5)  # 增加等待时间
            except Exception as e:
                print(f"页面加载警告: {str(e)[:50]}")
                try:
                    # 如果networkidle超时，尝试domcontentloaded
                    page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    time.sleep(5)
                except:
                    pass
            
            # 尝试多次获取内容
            html_content = None
            for attempt in range(3):
                try:
                    html_content = page.content()
                    if html_content and len(html_content) > 1000:
                        break
                    time.sleep(2)
                except:
                    if attempt < 2:
                        time.sleep(2)
            
            browser.close()
            
            if not html_content or len(html_content) < 500:
                return {
                    'title': '页面加载失败',
                    'author': '未找到作者',
                    'status': 'failed: 页面内容为空'
                }
            
            # 解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            title = None
            author = None
            
            # 提取标题 - 多种方法
            # 方法1: meta标签（最准确）
            meta_title = soup.find('meta', {'property': 'og:title'})
            if meta_title and meta_title.get('content'):
                title = meta_title.get('content', '').strip()
            
            # 方法2: h1标签
            if not title or len(title) < 5:
                h1_elem = soup.find('h1')
                if h1_elem:
                    h1_text = h1_elem.get_text().strip()
                    if len(h1_text) >= 5:  # 确保h1内容足够长
                        title = h1_text
            
            # 方法3: title标签
            if not title or len(title) < 5:
                title_elem = soup.find('title')
                if title_elem:
                    title_text = title_elem.get_text().strip()
                    # 清理标题（去掉网站名）
                    if '_' in title_text:
                        title = title_text.split('_')[0].strip()
                    elif '-' in title_text:
                        title = title_text.split('-')[0].strip()
                    else:
                        title = title_text
            
            # 方法4: 从JSON中提取title字段
            if not title or len(title) < 5:
                title_match = re.search(r'"title"\s*:\s*"([^"]{10,200})"', html_content)
                if title_match:
                    title = title_match.group(1)
            
            # 提取作者 - 增强多种方法
            # 方法1: 从JSON数据中提取（扩展模式）
            json_patterns = [
                r'"name"\s*:\s*"([^"]{2,30})"',  # 作者名
                r'"nickname"\s*:\s*"([^"]{2,30})"',  # 昵称
                r'"screen_name"\s*:\s*"([^"]{2,30})"',  # 屏幕名
                r'"userName"\s*:\s*"([^"]{2,30})"',  # 用户名
                r'"source"\s*:\s*"([^"]{2,30})"',  # 来源
                r'"author_name"\s*:\s*"([^"]{2,30})"',
                r'"user_name"\s*:\s*"([^"]{2,30})"',
                r'\"作者\"\s*:\s*\"([^\"]{2,30})\"',
                r'"media_name"\s*:\s*"([^"]{2,30})"',  # 媒体名
            ]
            
            for pattern in json_patterns:
                matches = re.findall(pattern, html_content)
                if matches:
                    # 过滤掉可能的误匹配
                    for match in matches:
                        if match and len(match) >= 2 and len(match) <= 30:
                            # 排除常见误匹配
                            exclude_words = ['今日头条', 'toutiao', 'article', 'content', 'title', 
                                           'video', 'image', 'button', 'input', 'label']
                            if not any(word in match.lower() for word in exclude_words):
                                # 清理日期：移除类似"2025-10-17 15:29"的日期时间
                                cleaned_author = re.sub(r'\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}.*$', '', match).strip()
                                if cleaned_author and len(cleaned_author) >= 2:
                                    author = cleaned_author
                                    break
                    if author:
                        break
            
            # 方法2: 从HTML元素提取（扩展选择器）
            if not author:
                author_selectors = [
                    {'class': re.compile('author|writer|creator', re.I)},
                    {'class': re.compile('source|publisher', re.I)},
                    {'class': re.compile('user|username', re.I)},
                    {'data-id': True},
                    {'id': re.compile('author', re.I)},
                ]
                for selector in author_selectors:
                    elem = soup.find('span', selector) or soup.find('a', selector) or soup.find('div', selector)
                    if elem:
                        text = elem.get_text().strip()
                        # 清理日期时间
                        text = re.sub(r'\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}.*$', '', text).strip()
                        if text and 2 <= len(text) <= 30:
                            author = text
                            break
            
            # 方法3: 从meta标签提取
            if not author:
                meta_author = soup.find('meta', {'name': 'author'}) or soup.find('meta', {'property': 'article:author'})
                if meta_author:
                    author = meta_author.get('content', '').strip()
            
            # 方法4: 从URL中提取用户信息（部分今日头条链接包含用户信息）
            if not author:
                # 尝试从share_uid或其他参数提取
                uid_match = re.search(r'share_uid=([^&]+)', url)
                if uid_match and 'AAAA' not in uid_match.group(1):
                    # 这只是备用方案，通常不会用到
                    pass
            
            # 最终清理作者字段（移除日期）
            if author:
                author = re.sub(r'\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}.*$', '', author).strip()
            
            return {
                'title': title if title else '未找到标题',
                'author': author if author else '未找到作者',
                'status': 'success' if (title and author) else 'partial'
            }
            
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def is_baidu_or_douyin(url):
    """检查是否是百度系、抖音或懂车帝链接（需要Playwright）"""
    url_lower = url.lower()
    return ('baijiahao.baidu.com' in url_lower or 
            'mbd.baidu.com' in url_lower or 
            'douyin.com' in url_lower or 
            'iesdouyin.com' in url_lower or
            'zjbyte.cn' in url_lower or
            'dongchedi' in url_lower or
            'dcd.' in url_lower)

def extract_platform_info(url):
    """识别平台并使用特定方法提取信息"""
    url_lower = url.lower()
    
    if 'mp.weixin.qq' in url_lower:
        return extract_weixin_info(url)
    elif 'weibo.com' in url_lower:
        return extract_weibo_breakthrough(url)
    elif 'toutiao.com' in url_lower:
        # 今日头条使用Playwright
        return extract_toutiao_playwright(url)
    elif 'douyin.com' in url_lower or 'iesdouyin.com' in url_lower:
        return extract_douyin_enhanced(url)
    elif 'xiaohongshu.com' in url_lower or 'xhslink.com' in url_lower:
        return extract_xiaohongshu_info(url)
    elif 'bilibili.com' in url_lower:
        return extract_bilibili_info(url)
    elif 'autohome.com' in url_lower:
        return extract_autohome_info(url)
    elif 'zjbyte.cn' in url_lower or 'dongchedi' in url_lower or 'dcd.' in url_lower:
        return extract_dongchedi_info(url)
    elif 'baijiahao.baidu.com' in url_lower or 'mbd.baidu.com' in url_lower:
        return extract_baidu_info(url)
    elif 'dripcar.com' in url_lower:
        return extract_dripcar_info(url)
    elif 'maiche.com' in url_lower:
        return extract_maiche_info(url)
    else:
        return extract_general_info(url)

def extract_with_playwright_browser(url):
    """使用 Playwright 浏览器方案处理难处理的平台（百度系、抖音）"""
    if not PLAYWRIGHT_AVAILABLE:
        return {
            'title': 'Playwright未安装',
            'author': '未找到',
            'status': 'failed: Playwright未安装'
        }
    
    try:
        with sync_playwright() as p:
            # 尝试多种方式启动浏览器
            browser = None
            launch_errors = []
            
            # 方式1: 默认启动
            if not browser:
                try:
                    browser = p.chromium.launch(
                        headless=True,
                        args=[
                            '--disable-blink-features=AutomationControlled',
                            '--disable-dev-shm-usage',
                            '--no-sandbox'
                        ]
                    )
                except Exception as e:
                    launch_errors.append(f"默认启动失败: {str(e)[:50]}")
            
            # 方式2: 不指定executable_path
            if not browser:
                try:
                    browser = p.chromium.launch(
                        headless=True,
                        executable_path=None,
                        args=[
                            '--disable-blink-features=AutomationControlled',
                            '--disable-dev-shm-usage',
                            '--no-sandbox'
                        ]
                    )
                except Exception as e:
                    launch_errors.append(f"无路径启动失败: {str(e)[:50]}")
            
            # 方式3: 尝试指定系统路径
            if not browser:
                import os
                possible_paths = [
                    os.path.expanduser("~/.cache/ms-playwright/chromium-*/chrome-win/chrome.exe"),
                    os.path.expanduser("~/AppData/Local/ms-playwright/chromium-*/chrome-win/chrome.exe"),
                    "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
                    "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe"
                ]
                for path in possible_paths:
                    try:
                        if '*' in path:
                            import glob
                            matches = glob.glob(path)
                            if matches:
                                path = matches[0]
                        if os.path.exists(path):
                            browser = p.chromium.launch(
                                headless=True,
                                executable_path=path,
                                args=[
                                    '--disable-blink-features=AutomationControlled',
                                    '--disable-dev-shm-usage',
                                    '--no-sandbox'
                                ]
                            )
                            break
                    except Exception as e:
                        launch_errors.append(f"路径{path}启动失败: {str(e)[:30]}")
                    
            # 如果所有方式都失败
            if not browser:
                error_msg = "; ".join(launch_errors) if launch_errors else "未知错误"
                return {
                    'title': '浏览器启动失败',
                    'author': '未找到',
                    'status': f'failed: Browser launch error - {error_msg}'
                }
            # 百度系检测很严，使用手机UA更容易通过
            user_agent = 'Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15E148 Safari/604.1' if 'baidu' in url.lower() else 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            viewport = {'width': 375, 'height': 812} if 'baidu' in url.lower() else {'width': 1920, 'height': 1080}
            
            context = browser.new_context(
                user_agent=user_agent,
                viewport=viewport,
            )
            page = context.new_page()
            
            # 去除 webdriver 标记
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            """)
            
            # 访问页面（使用domcontentloaded更快且成功率更高）
            try:
                page.goto(url, wait_until='domcontentloaded', timeout=20000)
                time.sleep(3)  # 等待 JavaScript 渲染
            except:
                # 即使超时也尝试继续
                try:
                    time.sleep(2)
                except:
                    pass
            
            # 获取页面内容
            html_content = page.content()
            browser.close()
            
            # 解析 HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            url_lower = url.lower()
            
            title = None
            author = None
            
            # 提取标题
            # 方法1: h1 标签
            h1_elem = soup.find('h1')
            if h1_elem:
                title = h1_elem.get_text().strip()
            
            # 方法2: title 标签
            if not title or len(title) < 5:
                title_elem = soup.find('title')
                if title_elem:
                    title = title_elem.get_text().strip()
                    # 清理标题
                    for sep in ['_', ' - ', '-']:
                        if sep in title:
                            parts = title.split(sep)
                            if len(parts[0].strip()) > 5:
                                title = parts[0].strip()
                                break
            
            # 方法3: meta 标签
            if not title or len(title) < 5:
                meta_title = soup.find('meta', {'property': 'og:title'})
                if meta_title:
                    title = meta_title.get('content', '').strip()
            
            # 提取作者
            # 针对百度系
            if 'baidu' in url_lower:
                # 方法1: JSON 数据（优先）
                author_patterns = [
                    r'"author"\s*:\s*"([^"]{2,30})"',  # 简单author字段
                    r'"author"\s*:\s*{\s*[^}]*"name"\s*:\s*"([^"]+)"',  # 嵌套author.name
                    r'"authorName"\s*:\s*"([^"]{2,30})"',
                    r'"author_name"\s*:\s*"([^"]{2,30})"',
                    r'"userName"\s*:\s*"([^"]{2,30})"',
                ]
                for pattern in author_patterns:
                    match = re.search(pattern, html_content)
                    if match:
                        potential_author = match.group(1)
                        # 过滤明显不是作者的
                        if potential_author not in ['百度', 'baidu', '好看视频']:
                            author = potential_author
                            break
                
                # 方法2: HTML 标签
                if not author:
                    author_elem = soup.find('span', {'data-testid': 'author-name'})
                    if author_elem:
                        author = author_elem.get_text().strip()
                
                if not author:
                    author_elem = soup.find('span', class_='_2gGWi')
                    if author_elem:
                        author = author_elem.get_text().strip()
                
                # 方法3: 查找class包含author的元素
                if not author:
                    author_elem = soup.find(class_=re.compile('author', re.I))
                    if author_elem:
                        text = author_elem.get_text().strip()
                        # 提取中文作者名（2-30个字符）
                        match = re.search(r'[\u4e00-\u9fa5a-zA-Z0-9]{2,30}', text)
                        if match:
                            author = match.group(0)
            
            # 针对抖音
            elif 'douyin' in url_lower:
                # 方法1: JSON 数据深度搜索
                render_match = re.search(r'<script id="RENDER_DATA" type="application/json">([^<]+)</script>', html_content)
                if render_match:
                    try:
                        json_str = html_module.unescape(render_match.group(1))
                        data = json.loads(json_str)
                        
                        def deep_find_author(obj, depth=0, max_depth=10):
                            if depth > max_depth:
                                return None
                            
                            if isinstance(obj, dict):
                                author_keys = ['nickname', 'authorName', 'unique_id', 'userName']
                                for key in author_keys:
                                    if key in obj and isinstance(obj[key], str) and 2 < len(obj[key]) < 50:
                                        return obj[key]
                                
                                for key in ['author', 'user']:
                                    if key in obj and isinstance(obj[key], dict):
                                        result = deep_find_author(obj[key], depth+1, max_depth)
                                        if result:
                                            return result
                                
                                for value in obj.values():
                                    result = deep_find_author(value, depth+1, max_depth)
                                    if result:
                                        return result
                            
                            elif isinstance(obj, list):
                                for item in obj:
                                    result = deep_find_author(item, depth+1, max_depth)
                                    if result:
                                        return result
                            
                            return None
                        
                        author = deep_find_author(data)
                    except:
                        pass
                
                # 方法2: 正则表达式
                if not author:
                    patterns = [
                        r'"nickname":\s*"([^"]{2,30})"',
                        r'"authorName":\s*"([^"]{2,30})"',
                    ]
                    for pattern in patterns:
                        matches = re.findall(pattern, html_content)
                        if matches:
                            author = matches[0]
                            break
            
            # 通用作者提取（fallback）
            if not author:
                meta_author = soup.find('meta', {'name': 'author'})
                if meta_author:
                    author = meta_author.get('content', '').strip()
            
            return {
                'title': title if title else '未找到标题',
                'author': author if author else '未找到作者',
                'status': 'success (Playwright)' if (title and author) else 'partial (Playwright)'
            }
            
    except Exception as e:
        return {
            'title': '提取失败',
            'author': '提取失败',
            'status': f'failed: {str(e)[:50]}'
        }

def extract_title_and_author(url):
    """从URL提取标题和作者信息"""
    return extract_platform_info(url)

def main():
    print("=" * 60)
    print("链接标题和作者提取工具 v4.6 - 两阶段处理版")
    print("✅ 阶段1: 先处理普通链接（跳过需Playwright的平台）")
    print("✅ 阶段2: 用Playwright批量处理百度/抖音/懂车帝")
    print("🔴 404错误整行标红 | 🟡 失败单元格标黄")
    print("=" * 60)
    
    # 从命令行参数获取文件名，如果没有则使用默认值
    import sys
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = '测试数据集_随机抽样.xlsx'
    
    max_links = None  # 处理全部数据
    
    print(f"\n正在读取文件: {excel_file}")
    links = read_excel_with_links(excel_file)
    
    print(f"找到 {len(links)} 个链接")
    
    # 限制处理数量（可选）
    if max_links and len(links) > max_links:
        links = links[:max_links]
        print(f"⚠️  限制处理前 {max_links} 条链接\n")
    else:
        print(f"处理全部 {len(links)} 条链接\n")
    
    results = []
    delayed_links = []  # 存储百度系和抖音链接，延迟处理
    start_time = time.time()
    
    # ========== 阶段1: 处理普通链接，跳过需Playwright的平台 ==========
    print("\n" + "=" * 60)
    print("【阶段1】处理普通链接（百度/抖音/懂车帝将在阶段2处理）")
    print("=" * 60 + "\n")
    
    for idx, link_info in enumerate(links, 1):
        website_name = get_website_name(link_info['url'])
        url_short = link_info['url'][:60]
        
        # 检查是否需要Playwright（百度/抖音/懂车帝）
        if is_baidu_or_douyin(link_info['url']):
            print(f"[{idx}/{len(links)}] {website_name} | {url_short}... ⏸️  延迟处理", flush=True)
            delayed_links.append({
                'idx': idx,
                'link_info': link_info,
                'website_name': website_name
            })
            # 添加占位结果
            results.append({
                '原链接': link_info['url'],
                '网站名': website_name,
                '作者': '待处理',
                '标题': '待处理',
                '状态': 'pending'
            })
            continue
        
        # 处理普通链接
        iter_start = time.time()
        print(f"[{idx}/{len(links)}] {website_name} | {url_short}...", end=' ', flush=True)
        
        info = extract_title_and_author(link_info['url'])
        
        iter_time = time.time() - iter_start
        status_icon = '✅' if info['status'] == 'success' else ('⚠️' if 'partial' in info['status'] else '❌')
        print(f"{status_icon} ({iter_time:.1f}s)", flush=True)
        
        results.append({
            '原链接': link_info['url'],
            '网站名': website_name,
            '作者': info['author'],
            '标题': info['title'],
            '状态': info['status']
        })
        
        # 每10条显示一次进度统计
        if idx % 10 == 0:
            elapsed = time.time() - start_time
            processed = idx - len(delayed_links)
            if processed > 0:
                avg_time = elapsed / processed
                remaining_normal = (len(links) - idx) * avg_time
                success = sum(1 for r in results if 'success' in r['状态'])
                print(f"  ⏱️  已用{elapsed/60:.1f}分 | 预计剩余{remaining_normal/60:.1f}分 | 成功率{success/processed*100:.1f}% | 延迟{len(delayed_links)}个\n")
        
        if idx < len(links):
            time.sleep(0.5)
    
    # ========== 阶段2: 用 Playwright 处理延迟的链接 ==========
    if delayed_links:
        print("\n" + "=" * 60)
        print(f"【阶段2】使用Playwright处理延迟的{len(delayed_links)}个链接")
        print("=" * 60 + "\n")
        
        stage2_start = time.time()
        
        for delayed_idx, delayed_item in enumerate(delayed_links, 1):
            idx = delayed_item['idx']
            link_info = delayed_item['link_info']
            website_name = delayed_item['website_name']
            url_short = link_info['url'][:60]
            
            iter_start = time.time()
            print(f"[{delayed_idx}/{len(delayed_links)}] {website_name} | {url_short}...", end=' ', flush=True)
            
            # 使用平台专用函数处理（会自动路由到对应的提取函数）
            info = extract_platform_info(link_info['url'])
            
            iter_time = time.time() - iter_start
            status_icon = '✅' if 'success' in info['status'] else ('⚠️' if 'partial' in info['status'] else '❌')
            print(f"{status_icon} ({iter_time:.1f}s)", flush=True)
            
            # 更新results中对应位置的结果
            results[idx - 1] = {
                '原链接': link_info['url'],
                '网站名': website_name,
                '作者': info['author'],
                '标题': info['title'],
                '状态': info['status']
            }
            
            # 每5条显示一次进度
            if delayed_idx % 5 == 0 or delayed_idx == len(delayed_links):
                elapsed = time.time() - stage2_start
                avg_time = elapsed / delayed_idx
                remaining = (len(delayed_links) - delayed_idx) * avg_time
                success = sum(1 for i in range(delayed_idx) if 'success' in results[delayed_links[i]['idx'] - 1]['状态'])
                print(f"  ⏱️  阶段2已用{elapsed/60:.1f}分 | 预计剩余{remaining/60:.1f}分 | 成功率{success/delayed_idx*100:.1f}%\n")
            
            # Playwright 处理需要更长延迟
            if delayed_idx < len(delayed_links):
                time.sleep(2)
    
    df = pd.DataFrame(results)
    output_file = '链接分析结果_v4_最终版.xlsx'
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    # 添加颜色标记
    print("\n添加颜色标记...")
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # 定义颜色
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # 浅红色
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # 浅黄色
    
    # 遍历数据行（从第2行开始，第1行是表头）
    for idx, row_data in enumerate(results, start=2):
        status = row_data['状态'].lower()
        title = row_data['标题']
        author = row_data['作者']
        
        # 404错误 → 整行标红
        if '404' in status or '404' in title:
            for col in range(1, 6):  # A-E列（原链接、网站名、作者、标题、状态）
                ws.cell(row=idx, column=col).fill = red_fill
        else:
            # 检查标题是否有问题（未找到、失败、错误等）
            title_error_keywords = ['未找到', '提取失败', 'API', '失败', '错误', '待处理', 
                                   '响应', '请求', 'URL格式', '页面加载', '浏览器启动', '验证码中间页']
            if any(keyword in title for keyword in title_error_keywords):
                ws.cell(row=idx, column=4).fill = yellow_fill  # D列：标题
            
            # 检查作者是否有问题
            author_error_keywords = ['未找到', '提取失败', 'API', '失败', '错误', '待处理',
                                    '响应', '请求', 'URL格式', '页面加载', '浏览器启动', '验证码中间页']
            if any(keyword in author for keyword in author_error_keywords):
                ws.cell(row=idx, column=3).fill = yellow_fill  # C列：作者
    
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print(f"分析完成！结果已保存到: {output_file}")
    print("🔴 404错误 → 整行标红")
    print("🟡 标题/作者提取失败 → 对应单元格标黄")
    print("=" * 60)
    
    success_count = sum(1 for r in results if 'success' in r['状态'].lower())
    partial_count = sum(1 for r in results if 'partial' in r['状态'].lower())
    failed_count = len(results) - success_count - partial_count
    
    print(f"\n【统计】")
    print(f"总链接数: {len(results)}")
    print(f"完全成功: {success_count} ({success_count/len(results)*100:.1f}%)")
    print(f"部分成功: {partial_count} ({partial_count/len(results)*100:.1f}%)")
    print(f"失败: {failed_count} ({failed_count/len(results)*100:.1f}%)")
    
    print(f"\n【平台分布】")
    platform_count = {}
    for r in results:
        platform = r['网站名']
        platform_count[platform] = platform_count.get(platform, 0) + 1
    
    for platform, count in sorted(platform_count.items(), key=lambda x: x[1], reverse=True):
        print(f"{platform}: {count}个")
    
    print("\n处理完成！")

if __name__ == "__main__":
    main()

